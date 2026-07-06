"""
Painel Mulheres — ETL da base "Vítimas de Violência Doméstica" (SSP, 2024+).

Fonte: assets/estatistica/vitimas-violencia-domestica/Base_Violência contra mulher.xlsx
(endpoint do portal SSP; 68 colunas, nível vítima×BO, inclui SEXO, IDADE,
RELACIONAMENTO com agressor, ORIENTAÇÃO SEXUAL, FLAG_VITIMA_FATAL).

Recorte: fato ocorrido na capital (NOME_MUNICIPIO_CIRC == S.PAULO — usar o município
de REGISTRO infla com o estado inteiro: a Delegacia da Mulher Online registra para
todo o estado e é sediada na capital), vítimas mulheres (SEXO_PESSOA == F),
tipos de pessoa vitimados (Vítima, Autor/Vítima, Criança, Adolescente*).

Saídas:
  - web/public/data/mulheres.bin   — microdado codificado: uint8, colunas concatenadas
  - web/public/data/mulheres_meta.json — dicionários, contagens e ordem das colunas
  - web/public/data/zonas.geojson  — 5 macrorregiões (subprefeituras dissolvidas) + pop

Zona do registro (cascata, documentada no rodapé do painel):
  1. coordenada → polígono da zona (spatial join)         (~41% dos registros)
  2. sem coordenada → seccional de polícia → zona          (resto; seccionais da
     capital são organizadas por zona: 1ª Centro ... 8ª São Mateus)

Agrupamento das naturezas pela tipologia da Lei Maria da Penha (art. 7º):
física, psicológica, moral, sexual, patrimonial (+ descumprimento de medida protetiva).
"""
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

import geopandas as gpd
import numpy as np
import pandas as pd
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
RAW = BASE / "data" / "raw" / "ViolenciaDomestica_Base.xlsx"
EXT = BASE / "data" / "external"
WEB = BASE / "web" / "public" / "data"

TIPOS_VITIMA = {"Vítima", "Autor/Vítima", "Criança", "Adolescente", "Adolescente Inf/Vit"}

ZONAS = ["Centro", "Norte", "Sul", "Leste", "Oeste"]
# As 8 seccionais da capital, pelo NÚMERO (a grafia º/ª varia na fonte e já causou
# bug: casar por texto deixava só o Centro com os registros sem coordenada):
# 1ª Centro, 2ª Sul, 3ª Oeste, 4ª Norte, 5ª Leste, 6ª Sto.Amaro→Sul,
# 7ª Itaquera→Leste, 8ª S.Mateus→Leste
SECCIONAL_NUM_ZONA = {1: 0, 2: 2, 3: 4, 4: 1, 5: 3, 6: 2, 7: 3, 8: 3}

SUB_ZONA = {  # 32 subprefeituras → 5 macrorregiões administrativas
    "SE": 0,
    "CASA VERDE": 1, "FREGUESIA DO O": 1, "FREGUESIA": 1, "JACANA": 1, "PERUS": 1,
    "PIRITUBA": 1, "SANTANA": 1, "VILA MARIA": 1,
    "CAMPO LIMPO": 2, "CAPELA DO SOCORRO": 2, "CIDADE ADEMAR": 2, "IPIRANGA": 2,
    "JABAQUARA": 2, "M BOI MIRIM": 2, "PARELHEIROS": 2, "SANTO AMARO": 2, "VILA MARIANA": 2,
    "ARICANDUVA": 3, "CIDADE TIRADENTES": 3, "ERMELINO MATARAZZO": 3, "GUAIANASES": 3,
    "ITAIM PAULISTA": 3, "ITAQUERA": 3, "MOOCA": 3, "PENHA": 3, "SAO MATEUS": 3,
    "SAO MIGUEL": 3, "SAPOPEMBA": 3, "VILA PRUDENTE": 3,
    "BUTANTA": 4, "LAPA": 4, "PINHEIROS": 4,
}

GRUPOS_MP = ["Física", "Psicológica", "Moral", "Sexual", "Patrimonial", "Medida protetiva"]
NATUREZA_GRUPO = {
    "LESÃO CORPORAL DOLOSA": 0, "MAUS TRATOS": 0,
    "AMEAÇA": 1, "PERSEGUIR": 1, "VIOLÊNCIA PSICOLÓGICA CONTRA A MULHER": 1,
    "CONSTRANGIMENTO ILEGAL": 1,
    "CALÚNIA - DIFAMAÇÃO - INJÚRIA": 2,
    "OUTROS C/C/ DIGNIDADE SEXUAL": 3, "DIVULGAÇÃO DE FOTOS/VIDEOS ÍNTIMOS": 3,
    "DANO": 4, "INVASÃO DE DOMICÍLIO": 4,
    "DESCUMPRIMENTO DE MEDIDA PROTETIVA DE URGÊNCIA": 5,
}

RELACOES = ["União estável", "Casamento", "Envolvimento amoroso", "Parentesco",
            "Convivência (amigo/vizinho/trabalho)", "Sem relação/outros", "Não informado"]
RELACAO_IDX = {
    "Uniao estavel": 0, "Casamento": 1, "Envolvimento amoroso": 2, "Parentesco": 3,
    "Conhecido": 4, "Amizade": 4, "Vizinhanca": 4, "Trabalho": 4,
    "Nenhuma relacao": 5, "Possivel env. em atividade criminosa": 5,
}

FAIXAS = ["0-11", "12-17", "18-24", "25-29", "30-39", "40-49", "50-59", "60+", "n/i"]

ORIENTACOES = ["Heterossexual", "Homossexual", "Bissexual", "Outras", "Não informado"]
ORIENTACAO_IDX = {"Heterossexual": 0, "Homossexual": 1, "Bissexual": 2, "Assexual": 3,
                  "Pansexual": 3, "Prefiro não me classificar: negativa de autodeclaração": 4}

LOCAIS = ["Residência", "Via pública", "Condomínio", "Internet",
          "Comércio/serviços", "Transporte/terminal", "Outros", "n/i"]
LOCAL_IDX = {
    "Residência": 0, "Condomínio Residencial": 2, "Internet": 3,
    "Via Pública": 1, "Comércio e Serviços": 4, "Restaurante e Afins": 4,
    "Terminal/Estação": 5, "Rodovia/Estrada": 5,
}

PERIODO_TXT = {"De madrugada": 0, "Pela manhã": 1, "A tarde": 2, "A noite": 3}
PERIODOS = ["madrugada", "manhã", "tarde", "noite", "n/i"]

# ordem das colunas no binário (uint8 cada)
COLUNAS_BIN = ["ano", "zona", "faixa", "relacao", "grupo", "natureza",
               "fatal", "orientacao", "periodo", "hora", "local"]


def norm(s: str) -> str:
    s = "".join(c for c in unicodedata.normalize("NFD", str(s)) if not unicodedata.combining(c))
    return s.upper().strip()


def seccional_para_zona(nome: str) -> int:
    m = re.search(r"DEL\.?\s*SEC\.?\s*(\d)", norm(nome or ""))
    return SECCIONAL_NUM_ZONA.get(int(m.group(1)), 5) if m else 5  # 5 = ignorada


def faixa_idade(v) -> int:
    if not isinstance(v, (int, float)) or v < 0 or v > 120:
        return 8
    for i, lim in enumerate([12, 18, 25, 30, 40, 50, 60]):
        if v < lim:
            return i
    return 7


def montar_zonas_geojson() -> gpd.GeoDataFrame:
    subs = gpd.read_file(WEB / "subprefeituras.geojson")
    # geojson traz nomes compostos ("CASA VERDE-LIMAO-CACHOEIRINHA") — casa pelo 1º componente
    subs["zona"] = subs["nome"].map(lambda n: SUB_ZONA.get(norm(n).split("-")[0].strip()))
    faltando = subs[subs["zona"].isna()]["nome"].tolist()
    assert not faltando, f"subprefeituras sem zona: {faltando}"
    zonas = subs.dissolve(by="zona", aggfunc={"pop_2022": "sum"}).reset_index()
    zonas["nome"] = zonas["zona"].map(lambda z: ZONAS[int(z)])
    zonas["zona"] = zonas["zona"].astype(int)
    zonas[["zona", "nome", "pop_2022", "geometry"]].to_file(WEB / "zonas.geojson", driver="GeoJSON")
    print(f"zonas.geojson: {len(zonas)} macrorregiões")
    return zonas


def main() -> None:
    zonas = montar_zonas_geojson()

    print("lendo base VD (262 MB, aguarde)...")
    wb = load_workbook(RAW, read_only=True, data_only=True)
    ws = wb["BASE_VD"]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    ix = {c: i for i, c in enumerate(header)}

    naturezas_vistas: dict[str, int] = {}
    serie_mensal: Counter[str] = Counter()  # "YYYY-MM" -> registros, p/ aba Séries
    linhas = []
    coords = []
    for r in it:
        if r[ix["NOME_MUNICIPIO_CIRC"]] != "S.PAULO":
            continue
        if r[ix["SEXO_PESSOA"]] != "F":
            continue
        if r[ix["DESCR_TIPO_PESSOA"]] not in TIPOS_VITIMA:
            continue
        nat = r[ix["NATUREZA_APURADA"]]
        if nat not in NATUREZA_GRUPO:
            continue

        ano = r[ix["ANO_ESTATISTICA"]]
        if ano not in (2024, 2025, 2026):
            continue

        mes = r[ix["MES_ESTATISTICA"]]
        if isinstance(mes, (int, float)) and 1 <= mes <= 12:
            serie_mensal[f"{int(ano)}-{int(mes):02d}"] += 1

        hora_raw = str(r[ix["HORA_OCORRENCIA_BO"]] or "")
        hora = 24
        if hora_raw[:2].strip().isdigit():
            h = int(hora_raw.split(":")[0])
            if 0 <= h <= 23:
                hora = h
        periodo = hora // 6 if hora < 24 else PERIODO_TXT.get(r[ix["DESCR_PERIODO"]], 4)

        lat, lon = r[ix["LATITUDE"]], r[ix["LONGITUDE"]]
        tem_coord = isinstance(lat, (int, float)) and isinstance(lon, (int, float)) and lat != 0
        coords.append((lon, lat) if tem_coord else None)

        rel = r[ix["DESCR_RELACIONAMENTO"]]
        idade = r[ix["IDADE_PESSOA"]]
        # zona: seccional do local do fato; se não for uma das 8 da capital
        # (fato fora/nulo), usa a seccional de registro como aproximação —
        # vítimas de violência doméstica costumam registrar perto de casa.
        zona = seccional_para_zona(r[ix["NOME_SECCIONAL_CIRC"]])
        if zona == 5:
            zona = seccional_para_zona(r[ix["NOME_SECCIONAL"]])
        linhas.append((
            int(ano) - 2024,
            zona,                                                # provisório; coord refina depois
            faixa_idade(idade),
            RELACAO_IDX.get(rel, 6),
            NATUREZA_GRUPO[nat],
            naturezas_vistas.setdefault(nat, len(naturezas_vistas)),
            1 if r[ix["FLAG_VITIMA_FATAL"]] == "S" else 0,
            ORIENTACAO_IDX.get(r[ix["DESCR_ORIENTACAO_SEXUAL"]], 4),
            int(periodo),
            hora,
            LOCAL_IDX.get(r[ix["DESCR_TIPOLOCAL"]], 7 if r[ix["DESCR_TIPOLOCAL"]] in (None, "NULL") else 6),
        ))

    df = pd.DataFrame(linhas, columns=COLUNAS_BIN).astype(np.uint8)
    print(f"{len(df):,} vítimas mulheres (capital, 2024-2026)")

    # refina a zona com a coordenada quando ela existe
    com = [(i, c) for i, c in enumerate(coords) if c]
    pontos = gpd.GeoDataFrame(
        {"i": [i for i, _ in com]},
        geometry=gpd.points_from_xy([c[0] for _, c in com], [c[1] for _, c in com]),
        crs=4326,
    )
    j = gpd.sjoin(pontos, zonas[["zona", "geometry"]], how="inner", predicate="within")
    j = j[~j.index.duplicated(keep="first")]
    df.loc[j["i"].to_numpy(), "zona"] = j["zona"].to_numpy().astype(np.uint8)
    via_coord = len(j)
    print(f"zona: {via_coord:,} via coordenada ({via_coord/len(df):.0%}); restante via seccional")

    WEB.mkdir(parents=True, exist_ok=True)
    (WEB / "mulheres.bin").write_bytes(
        b"".join(df[c].to_numpy().tobytes() for c in COLUNAS_BIN)
    )

    naturezas_ordem = [n for n, _ in sorted(naturezas_vistas.items(), key=lambda kv: kv[1])]
    meta = {
        "nrows": len(df),
        "colunas": COLUNAS_BIN,
        "rotulos": {
            "ano": ["2024", "2025", "2026"],
            "zona": ZONAS + ["n/i"],
            "faixa": FAIXAS,
            "relacao": RELACOES,
            "grupo": GRUPOS_MP,
            "natureza": [n.title() for n in naturezas_ordem],
            "fatal": ["Não fatal", "Fatal"],
            "orientacao": ORIENTACOES,
            "periodo": PERIODOS,
            "local": LOCAIS,
        },
        "cobertura_coord": round(via_coord / len(df), 3),
        "pop_zonas": {int(z): int(p) for z, p in zip(zonas["zona"], zonas["pop_2022"])},
        # total mensal (capital), independente do crossfilter — consumido pela aba Séries
        "serie_mensal": dict(sorted(serie_mensal.items())),
    }
    (WEB / "mulheres_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False), encoding="utf-8"
    )
    tam = (WEB / "mulheres.bin").stat().st_size
    print(f"mulheres.bin: {tam/1e6:.1f} MB | meta ok")


if __name__ == "__main__":
    main()
