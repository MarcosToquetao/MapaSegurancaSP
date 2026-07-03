"""
Fase 2 — Limpeza: xlsx SSP → parquet por ano (só capital, só categorias do MVP).

Regras validadas em docs/DICIONARIO.md:
  - filtro NOME_MUNICIPIO == 'S.PAULO'
  - NATUREZA_APURADA (auditada) → 4 categorias; demais naturezas ficam fora do MVP
  - 'NULL' literal → NA; lat/long 0 ou fora do bbox da capital → NA (fica só no coroplético)
  - leitura em streaming (openpyxl read_only) — os xlsx têm até ~600k linhas por aba

Uso:
    python pipeline/02_clean.py            # todos os xlsx presentes em data/raw/
    python pipeline/02_clean.py 2025       # ano específico
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
RAW_DIR = BASE / "data" / "raw"
OUT_DIR = BASE / "data" / "processed"

# bbox generoso do município de São Paulo (graus decimais)
LAT_MIN, LAT_MAX = -24.10, -23.30
LON_MIN, LON_MAX = -47.00, -46.30

CATEGORIA = {
    "HOMICÍDIO DOLOSO": "letais",
    "HOMICÍDIO DOLOSO POR ACIDENTE DE TRÂNSITO": "letais",
    "LATROCÍNIO": "letais",
    "LESÃO CORPORAL SEGUIDA DE MORTE": "letais",
    "ROUBO - OUTROS": "roubos",
    "ROUBO DE VEÍCULO": "roubos",
    "ROUBO DE CARGA": "roubos",
    "FURTO - OUTROS": "furtos",
    "FURTO DE VEÍCULO": "furtos",
    "FURTO DE CARGA": "furtos",
    "ESTUPRO": "genero",
    "ESTUPRO DE VULNERÁVEL": "genero",
    # base separada CelularesSubtraidos (naturezas derivadas da RUBRICA)
    "ROUBO DE CELULAR": "celular",
    "FURTO DE CELULAR": "celular",
}

COLUNAS = [
    "NUM_BO", "ANO_BO", "NOME_MUNICIPIO", "NATUREZA_APURADA", "RUBRICA",
    "DESCR_CONDUTA", "DESCR_TIPOLOCAL", "DESCR_SUBTIPOLOCAL", "BAIRRO",
    "LOGRADOURO", "NUMERO_LOGRADOURO", "LATITUDE", "LONGITUDE",
    "DATA_OCORRENCIA_BO", "HORA_OCORRENCIA_BO", "DESC_PERIODO",
    "MES_ESTATISTICA", "ANO_ESTATISTICA",
]


# o schema variou entre anos (2022 difere de 2025); normaliza nomes antes de indexar
ALIASES = {
    "CIDADE": "NOME_MUNICIPIO",
    "DESCR_PERIODO": "DESC_PERIODO",
}


def ler_aba(ws) -> pd.DataFrame | None:
    it = ws.iter_rows(values_only=True)
    header = [ALIASES.get(c, c) for c in next(it)]
    # aba-dicionário/capa (o nome dela varia entre anos): sem as colunas-chave, pula
    if "NOME_MUNICIPIO" not in header or "NATUREZA_APURADA" not in header:
        print(f"    (aba {ws.title!r} sem colunas de dados — ignorada)")
        return None
    pos = {c: header.index(c) for c in COLUNAS if c in header}
    ausentes = [c for c in COLUNAS if c not in pos]
    if ausentes:
        print(f"    (colunas ausentes nesta aba, preenchidas com NA: {ausentes})")
    imun, inat = header.index("NOME_MUNICIPIO"), header.index("NATUREZA_APURADA")
    linhas = [
        [r[pos[c]] if c in pos else None for c in COLUNAS]
        for r in it
        if r[imun] == "S.PAULO" and r[inat] in CATEGORIA
    ]
    return pd.DataFrame(linhas, columns=COLUNAS)


def limpar(df: pd.DataFrame) -> pd.DataFrame:
    df = df.drop(columns=["NOME_MUNICIPIO"])
    df["categoria"] = df["NATUREZA_APURADA"].map(CATEGORIA)

    for col in ("LATITUDE", "LONGITUDE"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    fora = (
        ~df["LATITUDE"].between(LAT_MIN, LAT_MAX)
        | ~df["LONGITUDE"].between(LON_MIN, LON_MAX)
    )
    df.loc[fora, ["LATITUDE", "LONGITUDE"]] = pd.NA

    df["DATA_OCORRENCIA_BO"] = pd.to_datetime(df["DATA_OCORRENCIA_BO"], errors="coerce")

    # campos vindos do xlsx podem misturar int/str na mesma coluna (ex.: BAIRRO "742")
    numericas = {"LATITUDE", "LONGITUDE", "ANO_BO", "MES_ESTATISTICA", "ANO_ESTATISTICA"}
    texto = [c for c in df.columns if c not in numericas and c != "DATA_OCORRENCIA_BO"]
    for c in texto:
        df[c] = df[c].astype("string").replace({"NULL": pd.NA, "": pd.NA, "None": pd.NA})
    return df


def processar(ano: int) -> None:
    origem = RAW_DIR / f"SPDadosCriminais_{ano}.xlsx"
    destino = OUT_DIR / f"ocorrencias_{ano}.parquet"
    print(f"[{ano}] lendo {origem.name} ...")
    wb = load_workbook(origem, read_only=True)
    partes = [p for s in wb.sheetnames if (p := ler_aba(wb[s])) is not None]
    df = limpar(pd.concat(partes, ignore_index=True))
    df.to_parquet(destino, index=False)
    geo = df["LATITUDE"].notna().mean()
    print(f"[{ano}] {len(df):,} ocorrências (capital, 4 categorias) | {geo:.0%} com coordenada -> {destino.name}")


def processar_celulares(ano: int) -> None:
    """
    Base CelularesSubtraidos: uma linha por OBJETO subtraído (um BO pode ter vários
    celulares). Contamos OCORRÊNCIAS: dedupe por (ID_DELEGACIA, ANO_BO, NUM_BO),
    mantendo a maior VERSAO do BO. Natureza derivada da RUBRICA:
    'Roubo (art. 157)...' → ROUBO DE CELULAR ; 'Furto (art. 155)...' → FURTO DE CELULAR.
    Filtros: CIDADE == S.PAULO (local do fato) e FLAG_STATUS == CONSUMADO.
    """
    origem = RAW_DIR / f"CelularesSubtraidos_{ano}.xlsx"
    destino = OUT_DIR / f"celulares_{ano}.parquet"
    print(f"[celular {ano}] lendo {origem.name} ...")
    wb = load_workbook(origem, read_only=True)

    CAMPOS = [
        "ID_DELEGACIA", "ANO_BO", "NUM_BO", "VERSAO", "CIDADE", "RUBRICA",
        "DESCR_CONDUTA", "DESCR_TIPOLOCAL", "DESCR_SUBTIPOLOCAL", "BAIRRO",
        "LOGRADOURO", "NUMERO_LOGRADOURO", "LATITUDE", "LONGITUDE",
        "DATA_OCORRENCIA_BO", "HORA_OCORRENCIA", "DESCR_PERIODO", "FLAG_STATUS",
        "MARCA_OBJETO", "MES_REGISTRO_BO", "ANO_REGISTRO_BO",
    ]
    # 2022 usa MES/ANO em vez de MES_REGISTRO_BO/ANO_REGISTRO_BO
    ALIASES_CEL = {"MES": "MES_REGISTRO_BO", "ANO": "ANO_REGISTRO_BO"}
    partes = []
    for s in wb.sheetnames:
        ws = wb[s]
        it = ws.iter_rows(values_only=True)
        header = [ALIASES_CEL.get(c, c) for c in next(it)]
        if "NUM_BO" not in header or "RUBRICA" not in header:
            continue
        pos = {c: header.index(c) for c in CAMPOS if c in header}
        icid, ist = header.index("CIDADE"), header.index("FLAG_STATUS")
        # strip: o arquivo de 2022 preenche CIDADE com espaços à direita
        linhas = [
            [r[pos[c]] if c in pos else None for c in CAMPOS]
            for r in it
            if str(r[icid]).strip() == "S.PAULO" and str(r[ist]).strip() == "CONSUMADO"
        ]
        partes.append(pd.DataFrame(linhas, columns=CAMPOS))
    df = pd.concat(partes, ignore_index=True)

    # objeto → ocorrência: mantém a versão mais recente de cada BO
    df = (df.sort_values("VERSAO", ascending=False)
            .drop_duplicates(["ID_DELEGACIA", "ANO_BO", "NUM_BO"]))

    df["categoria"] = "celular"
    df["NATUREZA_APURADA"] = df["RUBRICA"].astype(str).str.upper().map(
        lambda r: "ROUBO DE CELULAR" if r.startswith("ROUBO")
        else ("FURTO DE CELULAR" if r.startswith("FURTO") else None))
    df = df[df["NATUREZA_APURADA"].notna()]

    # alinha ao contrato da base principal (03/04/05 leem os dois formatos)
    df = df.rename(columns={
        "HORA_OCORRENCIA": "HORA_OCORRENCIA_BO",
        "DESCR_PERIODO": "DESC_PERIODO",
        "MES_REGISTRO_BO": "MES_ESTATISTICA",
        "ANO_REGISTRO_BO": "ANO_ESTATISTICA",
    }).drop(columns=["ID_DELEGACIA", "VERSAO", "CIDADE", "FLAG_STATUS"])
    df = limpar(df.assign(NOME_MUNICIPIO="S.PAULO"))
    df.to_parquet(destino, index=False)
    geo = df["LATITUDE"].notna().mean()
    print(f"[celular {ano}] {len(df):,} ocorrências (capital, consumadas) | "
          f"{geo:.0%} com coordenada -> {destino.name}")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    anos = [int(a) for a in sys.argv[1:]]
    if not anos:
        anos = sorted(
            int(p.stem.split("_")[1]) for p in RAW_DIR.glob("SPDadosCriminais_*.xlsx")
        )
    for ano in anos:
        processar(ano)
        if (RAW_DIR / f"CelularesSubtraidos_{ano}.xlsx").exists():
            processar_celulares(ano)


if __name__ == "__main__":
    main()
